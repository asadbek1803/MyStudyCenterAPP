import openpyxl
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from accounts.models import Account
from study_center.models import Fanlar, Group, StudyCenter
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

from django.core.exceptions import ObjectDoesNotExist


# Create your views here.

@login_required
def export_groups_to_excel(request):
    if request.user.is_authenticated and (request.user.role == 'director' or request.user.role == 'teacher'):
        # Excel fayl yaratish
        wb = Workbook()
        ws = wb.active
        ws.title = 'Groups'

        # Ustunlar nomlarini qo'shish
        columns = ['ID', 'Nomi', 'Guruh haqida', "Jami o'quvchi",  'URL manzili', 'Logo', 'Fan', "O'quv markaz", "Ustozlar"]
        ws.append(columns)

        # Ustun nomlarini formatlash
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Guruhlarni olish
        study_center = request.user.study_center
        groups = Group.objects.filter(study_center=study_center)
        
        if not groups.exists():
            return JsonResponse({"error": "Guruhlar topilmadi!"}, status=404)

        for group in groups:
            total_students = group.students.count()
            teachers_names = ', '.join(teacher.first_name for teacher in group.teachers.all())
            fan_names = ', '.join(fan.name for fan in group.fan.all())  # Fan nomlarini olish va birlashtirish
            ws.append([
                group.id,
                group.name,
                group.description,
                total_students,
                group.slug,
                group.group_logo.url if group.group_logo else '',
                fan_names,
                group.study_center.name if group.study_center else '',
                teachers_names
            ])

        # Excel maydonlarini formatlash
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(columns)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Ustun kengligini sozlash
        column_widths = [10, 30, 30, 15, 30, 20, 30, 30, 30]
        for i, column_width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width

        # Virtual fayl yaratish
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        # Excel faylni yaratish
        response = HttpResponse(file_stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="groups.xlsx"'

        return response
    else:
        return JsonResponse({"error": "Bu amalni o'qituvchi va direktorlar amalga oshira oladi!"}, status=500)
    
    
    
    

@login_required
def export_fan_to_excel(request):
    if request.user.is_authenticated and (request.user.role == 'director' or request.user.role == 'teacher'):
        # Excel fayl yaratish
        wb = Workbook()
        ws = wb.active
        ws.title = 'Fanlar'

        # Ustunlar nomlarini qo'shish
        columns = ['ID', 'Nomi', "Ustoz"]
        ws.append(columns)

        # Ustun nomlarini formatlash
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Fanlarni olish
        study_center = request.user.study_center
        fanlar = Fanlar.objects.filter(teacher__study_center=study_center)
        
        if not fanlar.exists():
            return JsonResponse({"error": "Fanlar topilmadi!"}, status=404)

        for fan in fanlar:
            teacher_name = fan.teacher.first_name if fan.teacher else 'Noma\'lum'
            ws.append([
                fan.id,
                fan.name,
                teacher_name
            ])

        # Excel maydonlarini formatlash
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(columns)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Ustun kengligini sozlash
        column_widths = [10, 30, 30]
        for i, column_width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width

        # Virtual fayl yaratish
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        # Excel faylni yaratish
        response = HttpResponse(file_stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="fanlar.xlsx"'

        return response
    else:
        return JsonResponse({"error": "Bu amalni o'qituvchi va direktorlar amalga oshira oladi!"}, status=404)
    
    
    
    
@login_required
def import_groups(request):
    if request.user.role == 'director' or request.user.role == 'teacher':
        if request.method == 'POST' and request.FILES['excel_file']:
            excel_file = request.FILES['excel_file']
            
            try:
                wb = load_workbook(excel_file)
                ws = wb.active

                for row in ws.iter_rows(min_row=2, values_only=True):
                    name = row[0] 
                    teachers = row[1] # Faraz qilamizki, birinchi ustun guruh nomi
                    description = row[2]
                    fan = row[3]
                    
                    # Boshqa ma'lumotlarni ham qo'shishingiz mumkin
                    
                    Group.objects.create(name=name, teachers=teachers, fan=fan, description=description).save()
                
                messages.success(request, "Guruhlar muvaffaqiyatli import qilindi")
            except Exception as e:
                messages.error(request, f"Xatolik yuz berdi: {str(e)}")
            
            return redirect('guruhlar')  # Guruhlar ro'yxati sahifasiga yo'naltirish
        
        return render(request, 'dashboard/import_guruh.html')
    else:
        return JsonResponse({"error": "Bu amalni o'qituvchi va direktorlar amalga oshira oladi!"}, status=404)
    

@login_required
def export_teachers_to_excel(request):
    if request.user.role in ['director', 'teacher']:
        # Excel fayl yaratish
        wb = Workbook()
        ws = wb.active
        ws.title = 'Ustozlar'
        
        # Ustunlar nomlarini qo'shish
        columns = ['ID', 'ISMI', 'FAMILYASI', 'Username', 'Email', 'Telefon raqami', 'Ustoz haqida', 'Guruhlari', 'Manzil', 'Oxirgi kirish', 'Ishlamoqdami']
        ws.append(columns)

        # Ustun nomlarini formatlash
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # O'qituvchilarni olish
        study_center = request.user.study_center
        teachers = study_center.teachers.all()
        
        if not teachers.exists():
            return JsonResponse({"error": "Ustozlar topilmadi!"}, status=404)

        for teacher in teachers:
            user_info = teacher.description if teacher.description else "Qo'shimcha ma'lumot yo'q"
            is_active = 'Ha' if teacher.is_active else 'Yo\'q'
            groups = ','.join(group.name for group in Group.objects.filter(teachers = teacher).all())
            ws.append([
                teacher.id,
                teacher.first_name,
                teacher.last_name,
                teacher.username,
                teacher.email,
                teacher.phone,
                user_info,
                groups,
                teacher.address,
                teacher.last_login.strftime('%d.%m.%Y %H:%M:%S'),
                is_active
            ])

        # Excel maydonlarini formatlash
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(columns)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Filtrlarni qo'shish
        ws.auto_filter.ref = ws.dimensions
        
        # Ustun kengligini sozlash
        column_widths = [10, 30, 30, 20, 30, 20, 30, 30, 20, 20]
        for i, column_width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width

        # Virtual fayl yaratish
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        # Excel faylni yaratish
        response = HttpResponse(file_stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="ustozlar.xlsx"'

        return response
    else:
        return JsonResponse({"error": "Bu amalni o'qituvchi va direktorlar amalga oshira oladi!"}, status=403)

@login_required
def export_students_excel(request):
    if request.user.role in ['director', 'teacher']:
        # Excel fayl yaratish
        wb = Workbook()
        ws = wb.active
        ws.title = 'Talabalar'
        
        # Ustunlar nomlarini qo'shish
        columns = ['ID', 'ISMI', 'FAMILYASI', 'Username', 'Email', 'Telefon raqami', 'Uy telefon raqami', 'Fanlari', 'Grandmi', 'Guruhi', 'Talaba haqida', 'Manzil', 'Jami balli', 'Oylik ball', 'Oxirgi kirish', "Hozir o'qiyaptimi"]
        ws.append(columns)
        
        # Ustun nomlarini formatlash
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Talabalarni olish
        study_center = request.user.study_center
        students = study_center.students.all()
        
        if not students.exists():
            return JsonResponse({"error": "Talabalar topilmadi!"}, status=404)
        
        for student in students:
            user_info = student.description if student.description else "O'quvchi haqida qo'shimcha ma'lumot yo'q"
            student_subjects = ', '.join(subject.name for subject in student.fanlar.all())
            groups = ', '.join(group.name for group in Group.objects.filter(students=student))
            grand = 'Ha' if student.is_grand else "Yo'q"
            study = 'Ha' if student.is_active else "Yo'q"
            ws.append([
                student.id,
                student.first_name,
                student.last_name,
                student.username,
                student.email,
                student.phone,
                student.phone_home,
                student_subjects,
                grand,
                groups,
                user_info,
                student.address,
                student.total_score,
                student.month_score,
                student.last_login.strftime('%d.%m.%Y %H:%M:%S'),
                study,
            ])
        
        # Excel maydonlarini formatlash
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(columns)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Filtrlarni qo'shish
        ws.auto_filter.ref = ws.dimensions
        
        # Ustun kengligini sozlash
        column_widths = [10, 30, 30, 20, 30, 20, 20, 30, 10, 20, 30, 30, 15, 15, 20, 20]
        for i, column_width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width

        # Virtual fayl yaratish
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        # Excel faylni yaratish
        response = HttpResponse(file_stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="talabalar.xlsx"'

        return response
    else:
        return JsonResponse({"error": "Bu amalni o'qituvchi va direktorlar amalga oshira oladi!"}, status=403)
    
    